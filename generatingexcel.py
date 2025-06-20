import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 1) 载入你现有的 Calculation + Input 工作簿
wb = load_workbook("PIAS_Showcase_Auto.xlsx")   # 改成你的文件名
if "ShowcaseSlide" in wb.sheetnames:
    wb.remove(wb["ShowcaseSlide"])
ws = wb.create_sheet("ShowcaseSlide")

# ---- 视觉常量 ----
PIAS_BLUE  = "005BAC"
PIAS_BROWN = "6F6149"
ACCENT_G   = "E2A53A"
GREY_BG    = "F7F7F7"
BORDER_T   = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

# ---- 基础排版 ----
ws.sheet_view.showGridLines = False
for i in range(1, 17):
    ws.column_dimensions[get_column_letter(i)].width = 11

# 2) 插入 Logo （占 A1:C5）
logo = XLImage("PIAS_Logo (1).png")          # 确认 Logo 路径
logo.height, logo.width = 110, 275          # 适配高度
ws.add_image(logo, "A1")

# 3) 标题
ws.merge_cells("E2:Q4")
ws["E2"].value = "Singlife Elite Term — Client Cash-Flow Snapshot"
ws["E2"].font  = Font(size=22, bold=True, color=PIAS_BROWN)
ws["E2"].alignment = Alignment(horizontal="left", vertical="center")

# 4) Milestone Age 下拉
ws["B7"] = "Milestone Age:"
ws["B7"].font = Font(bold=True, color=PIAS_BLUE)
ws["D7"] = 65
ws["D7"].alignment = Alignment(horizontal="center")
ws["D7"].border = BORDER_T
dv = DataValidation(type="list",
                    formula1='"55,60,65,70,75,80,85,90"',
                    allow_blank=False)
dv.add(ws["D7"]); ws.add_data_validation(dv)

# ----------------------------------------------------------
# 5) 深度数据（隐藏行 same as Auto 版本，略…）
#    Age / Premium / CumPrem / Guaranteed / Non-Gtd / Net / YieldSel
# ----------------------------------------------------------

# 6) 【左右 KPI 对比卡片】-----------------------------------
#    左：With Plan   右：Self-Insure
# ----------------------------------------------------------
# (1) 灰底矩形用合并单元格
ws.merge_cells("B10:E19")    # Plan
ws.merge_cells("F10:I19")    # Self

# (2) 卡片标题
ws["B10"].value = "With Plan"
ws["F10"].value = "If Self-Insure"
for cell in ("B10", "F10"):
    ws[cell].font = Font(size=14, bold=True, color=PIAS_BLUE)
    ws[cell].alignment = Alignment(horizontal="center")
    ws[cell].fill = PatternFill("solid", fgColor=GREY_BG)

# (3) KPI 文字 + 公式
plan_kpis = [("Total Premium (5y)",    "=SUM(B_DATA_RANGE)"),
             ("Break-Even Age",        "=..."),         # 同前面
             ("Yield @Sel",            "=INDEX(YIELDSEL_COL, MATCH(D7,AGE_COL,0))"),
             ("Maturity Benefit",      "=...")]

self_kpis = [("Lump-Sum Needed Today", "=Input!B9"),
             ("Monthly Deposit",       '=PMT(3%/12,25*12,-Input!B9)'),
             ("Total Out-of-Pocket",   "=Input!B9"),
             ("Opportunity Cost",      '=Input!B9*0.03*25')]

start = 12
for (lbl, f), (lbl2, f2) in zip(plan_kpis, self_kpis):
    ws[f"B{start}"] = lbl;  ws[f"F{start}"] = lbl2
    ws[f"E{start}"] = f    # Plan 数值
    ws[f"I{start}"] = f2   # Self 数值
    for col in ('B','C','D','E','F','G','H','I'):
        ws[f"{col}{start}"].border = BORDER_T
    ws[f"E{start}"].font = ws[f"I{start}"].font = Font(bold=True, color=PIAS_BROWN)
    ws[f"E{start}"].alignment = ws[f"I{start}"].alignment = Alignment(horizontal="right")
    start += 2

# ----------------------------------------------------------
# 7) 三张图：Premium 柱 + Cum vs Net 折 + Yield 散点
#    与上版 Auto 基本相同，这里略去代码
# ----------------------------------------------------------

# 8) 脚注
ws.merge_cells("B42:Q44")
ws["B42"].value = "All figures update with Input tab & milestone age (D7)."
ws["B42"].font  = Font(size=9, italic=True, color="666666")
ws["B42"].alignment = Alignment(horizontal="center", vertical="center")

# 9) 保存
wb.save("PIAS_Showcase_Final.xlsx")
